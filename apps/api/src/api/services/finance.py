"""複式簿記服務；所有餘額均由已過帳分錄計算。"""

from __future__ import annotations

import asyncio
import re
import uuid
from datetime import UTC, date, datetime
from decimal import ROUND_HALF_UP, Decimal

from fastapi import HTTPException
from sqlalchemy import exists, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from api.models.finance import (
    BudgetSubmissionStatus,
    ChartAccount,
    ExpenseClaimItem,
    ExpenseClaimItemEvidence,
    ExpenseClaimStatus,
    ExpensePaymentMethod,
    ExpensePaymentStatus,
    ExpenseProcurementStatus,
    FinanceAccountType,
    FinanceBudget,
    FinanceBudgetAllocation,
    FinanceBudgetAllocationRevision,
    FinanceBudgetNode,
    FinanceBudgetSubmission,
    FinanceLedger,
    FiscalPeriod,
    FundAccount,
    FundStorageType,
    JournalEntry,
    JournalLine,
    JournalStatus,
)
from api.models.org import Org
from api.schemas.finance import (
    BudgetAllocationCreate,
    BudgetAllocationUpdate,
    BudgetCreate,
    BudgetNodeCreate,
    BudgetReview,
    BudgetSubmissionCreate,
    ChartAccountUpdate,
    ExpenseBudgetUpdate,
    ExpenseClaimCreate,
    ExpenseProcurementUpdate,
    ExpenseReimbursementCreate,
    JournalCreate,
    TransferCreate,
)

DEFAULT_ACCOUNTS = (
    ("1101", "零用金", FinanceAccountType.ASSET),
    ("1102", "保險箱現金", FinanceAccountType.ASSET),
    ("1103", "銀行存款", FinanceAccountType.ASSET),
    ("1201", "應收款", FinanceAccountType.ASSET),
    ("2101", "應付款", FinanceAccountType.LIABILITY),
    ("3101", "累積餘絀", FinanceAccountType.EQUITY),
    ("4101", "活動收入", FinanceAccountType.REVENUE),
    ("4102", "商品收入", FinanceAccountType.REVENUE),
    ("4103", "學餐收入", FinanceAccountType.REVENUE),
    ("5101", "活動支出", FinanceAccountType.EXPENSE),
    ("5102", "行政支出", FinanceAccountType.EXPENSE),
    ("5103", "退款支出", FinanceAccountType.EXPENSE),
)

_EVIDENCE_KEY_RE = re.compile(r"^[0-9a-f]{32}\.(?:jpg|jpeg|png|webp|pdf)$")


def validate_evidence_key(evidence_key: str | None, ledger_id: uuid.UUID) -> None:
    """僅接受由本帳本私有上傳端點建立的儲存鍵，拒絕外部 URL 與跨帳本附件。"""
    if evidence_key is None:
        return
    prefix = f"finance/evidence/{ledger_id}/"
    filename = evidence_key.removeprefix(prefix)
    if (
        not evidence_key.startswith(prefix)
        or "/" in filename
        or not _EVIDENCE_KEY_RE.fullmatch(filename)
    ):
        raise HTTPException(400, "憑證必須透過本帳本的私有上傳端點取得")


def _claim_item_total(unit_price: int, tax_rate: int, quantity: Decimal | float) -> int:
    total = Decimal(unit_price) * (Decimal(100 + tax_rate) / Decimal(100)) * Decimal(quantity)
    return int(total.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


async def initialize_ledger(db: AsyncSession, org_id: uuid.UUID, name: str) -> FinanceLedger:
    existing = await db.scalar(select(FinanceLedger).where(FinanceLedger.org_id == org_id))
    if existing:
        return existing
    ledger = FinanceLedger(org_id=org_id, name=name)
    db.add(ledger)
    await db.flush()
    accounts: dict[str, ChartAccount] = {}
    for code, account_name, kind in DEFAULT_ACCOUNTS:
        account = ChartAccount(
            ledger_id=ledger.id, code=code, name=account_name, account_type=kind, is_system=True
        )
        accounts[code] = account
        db.add(account)
    await db.flush()
    for name_, storage, code in (
        ("零用金", FundStorageType.PETTY_CASH, "1101"),
        ("保險箱", FundStorageType.SAFE, "1102"),
        ("銀行帳戶", FundStorageType.BANK, "1103"),
    ):
        db.add(
            FundAccount(
                ledger_id=ledger.id,
                chart_account_id=accounts[code].id,
                name=name_,
                storage_type=storage,
            )
        )
    await db.flush()
    return ledger


async def get_ledger(db: AsyncSession, ledger_id: uuid.UUID) -> FinanceLedger:
    ledger = await db.get(FinanceLedger, ledger_id)
    if not ledger:
        raise HTTPException(404, "帳本不存在")
    return ledger


async def get_budget(db: AsyncSession, budget_id: uuid.UUID) -> FinanceBudget:
    budget = await db.get(FinanceBudget, budget_id)
    if not budget:
        raise HTTPException(404, "預算不存在")
    return budget


async def create_budget(
    db: AsyncSession, ledger_id: uuid.UUID, body: BudgetCreate
) -> FinanceBudget:
    period = await db.get(FiscalPeriod, body.period_id)
    if not period or period.ledger_id != ledger_id:
        raise HTTPException(400, "會計期間不屬於此帳本")
    existing = await db.scalar(
        select(FinanceBudget).where(
            FinanceBudget.ledger_id == ledger_id, FinanceBudget.period_id == body.period_id
        )
    )
    if existing:
        raise HTTPException(409, "此會計期間已有共同預算")
    budget = FinanceBudget(ledger_id=ledger_id, **body.model_dump())
    db.add(budget)
    await db.flush()
    return budget


async def list_budgets(db: AsyncSession, ledger_id: uuid.UUID) -> list[FinanceBudget]:
    return list(
        (
            await db.execute(
                select(FinanceBudget)
                .where(FinanceBudget.ledger_id == ledger_id)
                .order_by(FinanceBudget.created_at.desc())
            )
        ).scalars()
    )


def _import_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _import_decimal(value: object) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        parsed = Decimal(str(value).replace(",", "").strip())
    except (ValueError, ArithmeticError):
        return None
    return parsed if parsed.is_finite() and parsed > 0 else None


def _import_amount(value: object) -> int | None:
    parsed = _import_decimal(value)
    if parsed is None:
        return None
    return int(parsed.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _import_quantity(value: object) -> tuple[Decimal, str]:
    text = _import_text(value)
    if not text:
        return Decimal("1"), "項"
    match = re.match(r"^([0-9]+(?:\.[0-9]+)?)\s*(.*)$", text)
    if not match:
        return Decimal("1"), text[:32] or "項"
    quantity = Decimal(match.group(1))
    return quantity, match.group(2).strip()[:32] or "項"


def _parse_budget_workbook(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    from io import BytesIO

    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("無法讀取 xlsx 預算檔，請確認檔案未損壞") from exc

    try:
        rows = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()

    header_index = None
    columns: dict[str, int] = {}
    for index, row in enumerate(rows[:30]):
        normalized = {
            _import_text(value).replace(" ", ""): column for column, value in enumerate(row)
        }
        if "項目" in normalized and "細項" in normalized:
            header_index = index
            columns = normalized
            break
    if header_index is None:
        raise ValueError("找不到預算表欄位；至少需要「項目」與「細項」欄")

    category_column = columns["項目"]
    detail_column = columns["細項"]
    quantity_column = columns.get("數量", 3)
    unit_price_column = columns.get("單價", 4)
    amount_column = columns.get("總額(含稅)", columns.get("總額", 5))
    note_column = columns.get("備註", 7)
    imported: list[dict] = []
    skipped: list[str] = []
    current_category = ""
    for row_number, row in enumerate(rows[header_index + 1 :], header_index + 2):
        category = _import_text(row[category_column] if category_column < len(row) else None)
        detail = _import_text(row[detail_column] if detail_column < len(row) else None)
        if category:
            current_category = category
        if not detail:
            continue
        category = category or current_category or "未分類"
        quantity, unit = _import_quantity(
            row[quantity_column] if quantity_column < len(row) else None
        )
        unit_price = _import_amount(
            row[unit_price_column] if unit_price_column < len(row) else None
        )
        amount = _import_amount(row[amount_column] if amount_column < len(row) else None)
        if amount is None and unit_price is not None:
            amount = int((quantity * unit_price).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        if amount is None:
            skipped.append(f"第 {row_number} 列「{detail}」沒有可用總額")
            continue
        note = _import_text(row[note_column] if note_column < len(row) else None)
        if unit_price is not None:
            calculated = int((quantity * unit_price).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
            if calculated != amount:
                note = "；".join(filter(None, [note, "單價與總額不一致，保留試算表總額"]))
                unit_price = None
        elif _import_text(row[unit_price_column] if unit_price_column < len(row) else None):
            note = "；".join(filter(None, [note, "原始單價未提供，以試算表總額匯入"]))
        imported.append(
            {
                "category": category[:160],
                "detail": detail[:160],
                "quantity": quantity,
                "unit": unit,
                "unit_price": unit_price,
                "amount": amount,
                "note": note or None,
                "row_number": row_number,
            }
        )
    if not imported:
        raise ValueError("預算表沒有可匯入的明細")
    return imported, skipped


async def import_budget_from_xlsx(
    db: AsyncSession,
    ledger_id: uuid.UUID,
    period_id: uuid.UUID,
    name: str,
    title: str | None,
    file_bytes: bytes,
    user_id: uuid.UUID,
    proposing_org_id: uuid.UUID | None = None,
) -> tuple[FinanceBudget, FinanceBudgetSubmission, int, int, list[str]]:
    rows, skipped = await asyncio.to_thread(_parse_budget_workbook, file_bytes)
    ledger = await get_ledger(db, ledger_id)
    org_id = proposing_org_id or ledger.org_id
    if not await db.get(Org, org_id):
        raise HTTPException(400, "提出部門不存在")
    budget = await create_budget(db, ledger_id, BudgetCreate(period_id=period_id, name=name))
    submission = await create_budget_submission(
        db,
        budget,
        BudgetSubmissionCreate(
            kind="initial",
            title=(title or f"{name}（匯入）")[:160],
            note="由 xlsx 預算表匯入",
        ),
        user_id,
    )
    categories: dict[str, FinanceBudgetNode] = {}
    for sort_order, row in enumerate(rows):
        category = categories.get(row["category"])
        if category is None:
            category = await create_budget_node(
                db,
                submission.id,
                BudgetNodeCreate(name=row["category"], sort_order=sort_order),
            )
            categories[row["category"]] = category
        detail = await create_budget_node(
            db,
            submission.id,
            BudgetNodeCreate(parent_id=category.id, name=row["detail"], sort_order=sort_order),
        )
        await create_budget_allocation(
            db,
            submission.id,
            BudgetAllocationCreate(
                node_id=detail.id,
                amount=row["amount"],
                quantity=row["quantity"],
                unit=row["unit"],
                unit_price=row["unit_price"],
                proposing_org_id=org_id,
                note=row["note"],
            ),
            user_id,
        )
    return budget, submission, len(categories), len(rows), skipped


async def create_budget_submission(
    db: AsyncSession, budget: FinanceBudget, body: BudgetSubmissionCreate, user_id: uuid.UUID
) -> FinanceBudgetSubmission:
    if body.kind.value == "initial":
        existing = await db.scalar(
            select(FinanceBudgetSubmission.id).where(
                FinanceBudgetSubmission.budget_id == budget.id,
                FinanceBudgetSubmission.kind == body.kind,
            )
        )
        if existing:
            raise HTTPException(409, "共同預算只能建立一份初始預算案")
    submission = FinanceBudgetSubmission(
        budget_id=budget.id, created_by_id=user_id, **body.model_dump()
    )
    db.add(submission)
    await db.flush()
    return submission


async def _editable_submission(
    db: AsyncSession, submission_id: uuid.UUID
) -> FinanceBudgetSubmission:
    submission = await db.get(FinanceBudgetSubmission, submission_id)
    if not submission:
        raise HTTPException(404, "預算案不存在")
    if submission.status not in (BudgetSubmissionStatus.DRAFT, BudgetSubmissionStatus.RETURNED):
        raise HTTPException(400, "只有草案或退回的預算案可以編輯")
    return submission


async def create_budget_node(
    db: AsyncSession, submission_id: uuid.UUID, body: BudgetNodeCreate
) -> FinanceBudgetNode:
    submission = await _editable_submission(db, submission_id)
    budget = await get_budget(db, submission.budget_id)
    if body.parent_id:
        parent = await db.get(FinanceBudgetNode, body.parent_id)
        if not parent or parent.budget_id != budget.id:
            raise HTTPException(400, "上層預算條目不屬於此預算")
    node = FinanceBudgetNode(budget_id=budget.id, **body.model_dump())
    db.add(node)
    await db.flush()
    return node


async def create_budget_allocation(
    db: AsyncSession,
    submission_id: uuid.UUID,
    body: BudgetAllocationCreate,
    user_id: uuid.UUID,
) -> FinanceBudgetAllocation:
    submission = await _editable_submission(db, submission_id)
    node = await db.get(FinanceBudgetNode, body.node_id)
    if not node or node.budget_id != submission.budget_id:
        raise HTTPException(400, "預算條目不屬於此預算案")
    child = await db.scalar(
        select(FinanceBudgetNode.id).where(FinanceBudgetNode.parent_id == node.id).limit(1)
    )
    if child:
        raise HTTPException(400, "只有最末層預算條目可以配置金額")
    allocation = FinanceBudgetAllocation(
        submission_id=submission.id, proposed_by_id=user_id, **body.model_dump()
    )
    db.add(allocation)
    await db.flush()
    return allocation


async def update_budget_draft_allocation(
    db: AsyncSession,
    submission_id: uuid.UUID,
    allocation_id: uuid.UUID,
    body: BudgetAllocationCreate,
    user_id: uuid.UUID,
) -> FinanceBudgetAllocation:
    submission = await _editable_submission(db, submission_id)
    allocation = await db.get(FinanceBudgetAllocation, allocation_id)
    if not allocation or allocation.submission_id != submission.id:
        raise HTTPException(404, "預算配置不存在")
    if allocation.proposed_by_id != user_id:
        raise HTTPException(403, "只能修改自己提出的預算配置")
    node = await db.get(FinanceBudgetNode, body.node_id)
    if not node or node.budget_id != submission.budget_id:
        raise HTTPException(400, "預算條目不屬於此預算案")
    child = await db.scalar(
        select(FinanceBudgetNode.id).where(FinanceBudgetNode.parent_id == node.id).limit(1)
    )
    if child:
        raise HTTPException(400, "只有最末層預算條目可以配置金額")
    allocation.node_id = body.node_id
    assert body.amount is not None
    allocation.amount = body.amount
    allocation.quantity = body.quantity
    allocation.unit = body.unit
    allocation.unit_price = body.unit_price
    allocation.proposing_org_id = body.proposing_org_id
    allocation.note = body.note
    await db.flush()
    return allocation


async def submit_budget_submission(
    db: AsyncSession, submission_id: uuid.UUID
) -> FinanceBudgetSubmission:
    submission = await _editable_submission(db, submission_id)
    count = await db.scalar(
        select(func.count())
        .select_from(FinanceBudgetAllocation)
        .where(FinanceBudgetAllocation.submission_id == submission.id)
    )
    if not count:
        raise HTTPException(400, "預算案至少需要一筆最末層配置")
    submission.status = BudgetSubmissionStatus.SUBMITTED
    submission.submitted_at = datetime.now(UTC)
    await db.flush()
    return submission


async def review_budget_submission(
    db: AsyncSession, submission_id: uuid.UUID, body: BudgetReview, reviewer_id: uuid.UUID
) -> FinanceBudgetSubmission:
    submission = await db.get(FinanceBudgetSubmission, submission_id)
    if not submission:
        raise HTTPException(404, "預算案不存在")
    if submission.status != BudgetSubmissionStatus.SUBMITTED:
        raise HTTPException(400, "只有待審預算案可以審核")
    submission.status = body.status
    submission.review_note = body.note
    submission.reviewed_by_id = reviewer_id
    submission.reviewed_at = datetime.now(UTC)
    await db.flush()
    return submission


async def update_budget_allocation(
    db: AsyncSession,
    allocation_id: uuid.UUID,
    body: BudgetAllocationUpdate,
    user_id: uuid.UUID,
) -> FinanceBudgetAllocation:
    allocation = await db.get(FinanceBudgetAllocation, allocation_id)
    if not allocation:
        raise HTTPException(404, "預算配置不存在")
    submission = await db.get(FinanceBudgetSubmission, allocation.submission_id)
    if not submission or submission.status != BudgetSubmissionStatus.APPROVED:
        raise HTTPException(400, "只有已核准預算可以直接修正")
    db.add(
        FinanceBudgetAllocationRevision(
            allocation_id=allocation.id,
            previous_amount=allocation.amount,
            next_amount=body.amount,
            reason=body.reason,
            changed_by_id=user_id,
        )
    )
    allocation.amount = body.amount
    await db.flush()
    return allocation


async def budget_detail(db: AsyncSession, budget: FinanceBudget) -> dict:
    submissions = list(
        (
            await db.execute(
                select(FinanceBudgetSubmission)
                .where(FinanceBudgetSubmission.budget_id == budget.id)
                .order_by(FinanceBudgetSubmission.created_at)
            )
        ).scalars()
    )
    nodes = list(
        (
            await db.execute(
                select(FinanceBudgetNode)
                .where(FinanceBudgetNode.budget_id == budget.id, FinanceBudgetNode.is_active)
                .order_by(FinanceBudgetNode.sort_order, FinanceBudgetNode.name)
            )
        ).scalars()
    )
    allocations = list(
        (
            await db.execute(
                select(FinanceBudgetAllocation)
                .join(FinanceBudgetSubmission)
                .where(FinanceBudgetSubmission.budget_id == budget.id)
            )
        ).scalars()
    )
    approved_ids = {
        item.id for item in submissions if item.status == BudgetSubmissionStatus.APPROVED
    }
    totals: dict[uuid.UUID, int] = {}
    for allocation in allocations:
        if allocation.submission_id in approved_ids:
            totals[allocation.node_id] = totals.get(allocation.node_id, 0) + allocation.amount
    used_rows = (
        await db.execute(
            select(
                ExpenseClaimItem.budget_node_id,
                func.coalesce(
                    func.sum(
                        func.round(
                            ExpenseClaimItem.unit_price
                            * ((100 + ExpenseClaimItem.tax_rate) / 100)
                            * ExpenseClaimItem.quantity
                        )
                    ),
                    0,
                ),
            )
            .join(JournalEntry, JournalEntry.id == ExpenseClaimItem.journal_entry_id)
            .where(
                ExpenseClaimItem.budget_node_id.is_not(None),
                JournalEntry.status == JournalStatus.POSTED,
                JournalEntry.ledger_id == budget.ledger_id,
            )
            .group_by(ExpenseClaimItem.budget_node_id)
        )
    ).all()
    used = {row[0]: int(row[1]) for row in used_rows if row[0] is not None}
    return {
        "id": budget.id,
        "ledger_id": budget.ledger_id,
        "period_id": budget.period_id,
        "name": budget.name,
        "is_public": budget.is_public,
        "submissions": submissions,
        "allocations": allocations,
        "nodes": [
            {
                "id": node.id,
                "budget_id": node.budget_id,
                "parent_id": node.parent_id,
                "name": node.name,
                "sort_order": node.sort_order,
                "allocated_amount": totals.get(node.id, 0),
                "used_amount": used.get(node.id, 0),
                "remaining_amount": totals.get(node.id, 0) - used.get(node.id, 0),
            }
            for node in nodes
        ],
    }


async def set_budget_publication(
    db: AsyncSession, budget: FinanceBudget, is_public: bool
) -> FinanceBudget:
    if is_public:
        approved_initial = await db.scalar(
            select(FinanceBudgetSubmission.id).where(
                FinanceBudgetSubmission.budget_id == budget.id,
                FinanceBudgetSubmission.kind == "initial",
                FinanceBudgetSubmission.status == BudgetSubmissionStatus.APPROVED,
            )
        )
        if not approved_initial:
            raise HTTPException(400, "初始預算案核准後才能開放議員檢視")
    budget.is_public = is_public
    await db.flush()
    return budget


async def list_public_budgets(db: AsyncSession) -> list[tuple[FinanceBudget, FiscalPeriod]]:
    approved_initial = exists().where(
        FinanceBudgetSubmission.budget_id == FinanceBudget.id,
        FinanceBudgetSubmission.kind == "initial",
        FinanceBudgetSubmission.status == BudgetSubmissionStatus.APPROVED,
    )
    return list(
        (
            await db.execute(
                select(FinanceBudget, FiscalPeriod)
                .join(FiscalPeriod, FiscalPeriod.id == FinanceBudget.period_id)
                .where(FinanceBudget.is_public, approved_initial)
                .order_by(FiscalPeriod.starts_on.desc(), FinanceBudget.name)
            )
        ).all()
    )


async def public_budget_detail(db: AsyncSession, budget_id: uuid.UUID) -> tuple[dict, FiscalPeriod]:
    budget = await db.get(FinanceBudget, budget_id)
    if not budget or not budget.is_public:
        raise HTTPException(404, "公開預算不存在")
    approved_initial = await db.scalar(
        select(FinanceBudgetSubmission.id).where(
            FinanceBudgetSubmission.budget_id == budget.id,
            FinanceBudgetSubmission.kind == "initial",
            FinanceBudgetSubmission.status == BudgetSubmissionStatus.APPROVED,
        )
    )
    if not approved_initial:
        raise HTTPException(404, "公開預算不存在")
    period = await db.get(FiscalPeriod, budget.period_id)
    if not period:
        raise HTTPException(404, "會計期間不存在")
    return await budget_detail(db, budget), period


async def settlement_report(db: AsyncSession, ledger_id: uuid.UUID, period_id: uuid.UUID) -> dict:
    period = await db.get(FiscalPeriod, period_id)
    if not period or period.ledger_id != ledger_id:
        raise HTTPException(404, "會計期間不存在")
    budget = await db.scalar(
        select(FinanceBudget).where(
            FinanceBudget.ledger_id == ledger_id, FinanceBudget.period_id == period_id
        )
    )
    if not budget:
        return {
            "period_id": period.id,
            "period_name": period.name,
            "budgeted_total": 0,
            "settled_total": 0,
            "unsettled_claim_count": 0,
            "lines": [],
        }
    detail = await budget_detail(db, budget)
    settled_rows = (
        await db.execute(
            select(
                ExpenseClaimItem.budget_node_id,
                func.coalesce(
                    func.sum(
                        func.round(
                            ExpenseClaimItem.unit_price
                            * ((100 + ExpenseClaimItem.tax_rate) / 100)
                            * ExpenseClaimItem.quantity
                        )
                    ),
                    0,
                ),
            )
            .join(JournalEntry, JournalEntry.id == ExpenseClaimItem.journal_entry_id)
            .where(
                ExpenseClaimItem.budget_node_id.is_not(None),
                JournalEntry.ledger_id == ledger_id,
                JournalEntry.period_id == period_id,
                JournalEntry.status == JournalStatus.POSTED,
                JournalEntry.claim_status == ExpenseClaimStatus.COMPLETED,
            )
            .group_by(ExpenseClaimItem.budget_node_id)
        )
    ).all()
    settled = {row[0]: int(row[1]) for row in settled_rows if row[0] is not None}
    unsettled_claim_count = await db.scalar(
        select(func.count())
        .select_from(JournalEntry)
        .where(
            JournalEntry.ledger_id == ledger_id,
            JournalEntry.period_id == period_id,
            JournalEntry.source_type == "expense_claim",
            JournalEntry.status == JournalStatus.POSTED,
            JournalEntry.claim_status != ExpenseClaimStatus.COMPLETED,
        )
    )
    lines = [
        {
            "node_id": node["id"],
            "name": node["name"],
            "budgeted_amount": node["allocated_amount"],
            "settled_amount": settled.get(node["id"], 0),
            "difference_amount": node["allocated_amount"] - settled.get(node["id"], 0),
        }
        for node in detail["nodes"]
        if not any(child["parent_id"] == node["id"] for child in detail["nodes"])
    ]
    return {
        "period_id": period.id,
        "period_name": period.name,
        "budgeted_total": sum(line["budgeted_amount"] for line in lines),
        "settled_total": sum(line["settled_amount"] for line in lines),
        "unsettled_claim_count": int(unsettled_claim_count or 0),
        "lines": lines,
    }


async def update_chart_account(
    db: AsyncSession,
    ledger_id: uuid.UUID,
    account_id: uuid.UUID,
    body: ChartAccountUpdate,
) -> ChartAccount:
    account = await db.get(ChartAccount, account_id)
    if not account or account.ledger_id != ledger_id:
        raise HTTPException(404, "會計科目不存在")
    if body.name is not None:
        account.name = body.name
    if body.is_active is not None:
        if account.is_system and not body.is_active:
            raise HTTPException(400, "系統預設科目不可停用")
        account.is_active = body.is_active
    await db.flush()
    return account


async def validate_period(
    db: AsyncSession, ledger_id: uuid.UUID, period_id: uuid.UUID, entry_date: date
) -> FiscalPeriod:
    period = await db.get(FiscalPeriod, period_id)
    if not period or period.ledger_id != ledger_id:
        raise HTTPException(400, "會計期間不屬於此帳本")
    if period.is_closed:
        raise HTTPException(400, "會計期間已關閉")
    if not period.starts_on <= entry_date <= period.ends_on:
        raise HTTPException(400, "分錄日期不在會計期間內")
    return period


async def create_journal(
    db: AsyncSession,
    ledger_id: uuid.UUID,
    body: JournalCreate,
    user_id: uuid.UUID,
    *,
    pending: bool = False,
) -> JournalEntry:
    await validate_period(db, ledger_id, body.period_id, body.entry_date)
    validate_evidence_key(body.evidence_url, ledger_id)
    if body.source_type == "council_proposal":
        from api.models.council_proposal import CouncilProposal

        if not body.source_id or not await db.get(CouncilProposal, body.source_id):
            raise HTTPException(400, "議會提案關聯不存在")
    account_ids = {line.account_id for line in body.lines}
    count = await db.scalar(
        select(func.count())
        .select_from(ChartAccount)
        .where(
            ChartAccount.ledger_id == ledger_id,
            ChartAccount.id.in_(account_ids),
            ChartAccount.is_active,
        )
    )  # noqa: E712
    if count != len(account_ids):
        raise HTTPException(400, "分錄包含無效或停用科目")
    entry = JournalEntry(
        ledger_id=ledger_id,
        created_by_id=user_id,
        status=JournalStatus.PENDING_REVIEW if pending else JournalStatus.DRAFT,
        **body.model_dump(exclude={"lines"}),
    )
    db.add(entry)
    await db.flush()
    db.add_all([JournalLine(entry_id=entry.id, **line.model_dump()) for line in body.lines])
    await db.flush()
    return entry


async def submit_journal(db: AsyncSession, entry: JournalEntry) -> JournalEntry:
    if entry.status not in (JournalStatus.DRAFT, JournalStatus.RETURNED):
        raise HTTPException(400, "此傳票無法送覆核")
    entry.status = JournalStatus.PENDING_REVIEW
    if entry.source_type == "expense_claim":
        entry.claim_status = ExpenseClaimStatus.PENDING_REVIEW
    await db.flush()
    return entry


async def post_journal(
    db: AsyncSession, entry: JournalEntry, reviewer_id: uuid.UUID
) -> JournalEntry:
    # 財務過帳是不可逆的狀態轉換；鎖定傳票避免兩個覆核者同時通過檢查後重複過帳。
    locked_entry = await db.scalar(
        select(JournalEntry).where(JournalEntry.id == entry.id).with_for_update()
    )
    if locked_entry is None:
        raise HTTPException(404, "傳票不存在")
    entry = locked_entry
    if entry.status != JournalStatus.PENDING_REVIEW:
        raise HTTPException(400, "僅待覆核傳票可過帳")
    await validate_period(db, entry.ledger_id, entry.period_id, entry.entry_date)
    lines = list(
        (await db.execute(select(JournalLine).where(JournalLine.entry_id == entry.id))).scalars()
    )
    if len(lines) < 2 or sum(x.debit for x in lines) != sum(x.credit for x in lines):
        raise HTTPException(400, "傳票借貸不平衡")
    entry.status = JournalStatus.POSTED
    entry.reviewed_by_id = reviewer_id
    entry.posted_at = datetime.now(UTC)
    if entry.source_type == "expense_claim":
        entry.claim_status = ExpenseClaimStatus.APPROVED
    await db.flush()
    return entry


async def return_expense_claim(
    db: AsyncSession, entry: JournalEntry, reviewer_id: uuid.UUID, note: str
) -> JournalEntry:
    if entry.source_type != "expense_claim":
        raise HTTPException(400, "只有報帳案件可以退回")
    if entry.status != JournalStatus.PENDING_REVIEW:
        raise HTTPException(400, "僅待覆核報帳可以退回")
    entry.status = JournalStatus.RETURNED
    entry.claim_status = ExpenseClaimStatus.RETURNED
    entry.reviewed_by_id = reviewer_id
    entry.note = note
    await db.flush()
    return entry


async def update_expense_procurement(
    db: AsyncSession,
    entry: JournalEntry,
    body: ExpenseProcurementUpdate,
    user_id: uuid.UUID,
) -> JournalEntry:
    if entry.source_type != "expense_claim":
        raise HTTPException(400, "只有報帳案件可以管理校商請購")
    if entry.claim_status != ExpenseClaimStatus.APPROVED:
        raise HTTPException(400, "報帳必須先完成覆核")
    if entry.budget_included is not True:
        raise HTTPException(400, "報帳必須先列入已核准預算")
    current = ExpenseProcurementStatus(
        entry.procurement_status or ExpenseProcurementStatus.NOT_REQUIRED
    )
    allowed = {
        ExpenseProcurementStatus.NOT_REQUIRED: {
            ExpenseProcurementStatus.NOT_REQUIRED,
            ExpenseProcurementStatus.REQUESTED,
        },
        ExpenseProcurementStatus.REQUESTED: {
            ExpenseProcurementStatus.REQUESTED,
            ExpenseProcurementStatus.ORDERED,
            ExpenseProcurementStatus.NOT_REQUIRED,
        },
        ExpenseProcurementStatus.ORDERED: {
            ExpenseProcurementStatus.ORDERED,
            ExpenseProcurementStatus.RECEIVED,
        },
        ExpenseProcurementStatus.RECEIVED: {ExpenseProcurementStatus.RECEIVED},
    }
    if body.status not in allowed[current]:
        raise HTTPException(400, "校商請購狀態不可往回跳轉")
    entry.procurement_status = body.status
    entry.procurement_updated_by_id = user_id
    entry.procurement_updated_at = datetime.now(UTC)
    if body.note:
        entry.note = body.note
    await db.flush()
    return entry


async def mark_expense_paid(
    db: AsyncSession, entry: JournalEntry, payment_status: ExpensePaymentStatus, user_id: uuid.UUID
) -> JournalEntry:
    if entry.source_type != "expense_claim":
        raise HTTPException(400, "只有報帳案件可以登錄付款")
    if entry.claim_status != ExpenseClaimStatus.APPROVED or entry.status != JournalStatus.POSTED:
        raise HTTPException(400, "報帳必須先完成覆核並過帳")
    if entry.budget_included is not True:
        raise HTTPException(400, "報帳必須先列入已核准預算")
    if entry.payment_status not in (None, ExpensePaymentStatus.UNPAID):
        raise HTTPException(400, "此報帳已登錄付款，不可重複付款")
    entry.payment_status = payment_status
    entry.payment_by_id = user_id
    entry.payment_at = datetime.now(UTC)
    await db.flush()
    return entry


async def reimburse_expense_claim(
    db: AsyncSession,
    entry: JournalEntry,
    body: ExpenseReimbursementCreate,
    user_id: uuid.UUID,
) -> JournalEntry:
    if entry.source_type != "expense_claim" or entry.payment_method != ExpensePaymentMethod.ADVANCE:
        raise HTTPException(400, "只有代墊報帳可以建立償還付款")
    if entry.claim_status != ExpenseClaimStatus.APPROVED or entry.status != JournalStatus.POSTED:
        raise HTTPException(400, "代墊報帳必須先完成覆核並過帳")
    if entry.budget_included is not True:
        raise HTTPException(400, "代墊報帳必須先列入已核准預算")
    if entry.reimbursement_entry_id is not None:
        raise HTTPException(400, "此代墊報帳已完成償還")
    fund = await db.get(FundAccount, body.fund_account_id)
    if not fund or fund.ledger_id != entry.ledger_id or not fund.is_active:
        raise HTTPException(400, "付款資金保管點不存在或已停用")
    payable = await db.scalar(
        select(ChartAccount).where(
            ChartAccount.ledger_id == entry.ledger_id,
            ChartAccount.account_type == FinanceAccountType.LIABILITY,
            ChartAccount.code == "2101",
        )
    )
    if not payable:
        raise HTTPException(400, "找不到代墊應付款科目")
    amount = await db.scalar(
        select(func.coalesce(func.sum(JournalLine.credit), 0)).where(
            JournalLine.entry_id == entry.id, JournalLine.account_id == payable.id
        )
    )
    if not amount:
        raise HTTPException(400, "代墊報帳沒有可償還金額")
    reimbursement = await create_journal(
        db,
        entry.ledger_id,
        JournalCreate(
            period_id=body.period_id,
            entry_date=body.entry_date,
            description=f"代墊償還｜{entry.description}",
            source_type="expense_reimbursement",
            source_id=entry.id,
            source_event="reimbursement",
            note=body.note,
            lines=[
                {"account_id": payable.id, "debit": int(amount)},
                {"account_id": fund.chart_account_id, "credit": int(amount)},
            ],
        ),
        user_id,
    )
    reimbursement.status = JournalStatus.POSTED
    reimbursement.reviewed_by_id = user_id
    reimbursement.posted_at = datetime.now(UTC)
    entry.reimbursement_entry_id = reimbursement.id
    entry.payment_status = body.payment_status
    entry.payment_by_id = user_id
    entry.payment_at = datetime.now(UTC)
    await db.flush()
    return entry


async def update_expense_budget(
    db: AsyncSession,
    entry: JournalEntry,
    body: ExpenseBudgetUpdate,
    user_id: uuid.UUID,
) -> JournalEntry:
    if entry.source_type != "expense_claim":
        raise HTTPException(400, "只有報帳案件可以管理預算列管")
    if entry.claim_status != ExpenseClaimStatus.APPROVED:
        raise HTTPException(400, "報帳必須先完成覆核")
    if body.included:
        item_node_ids = set(
            (
                await db.execute(
                    select(ExpenseClaimItem.budget_node_id).where(
                        ExpenseClaimItem.journal_entry_id == entry.id
                    )
                )
            ).scalars()
        )
        if None in item_node_ids or not item_node_ids:
            raise HTTPException(400, "所有報帳品項都必須對應已核准的預算條目")
        approved_node_ids = set(
            (
                await db.execute(
                    select(FinanceBudgetAllocation.node_id)
                    .join(FinanceBudgetSubmission)
                    .where(
                        FinanceBudgetAllocation.node_id.in_(item_node_ids),
                        FinanceBudgetSubmission.status == BudgetSubmissionStatus.APPROVED,
                    )
                )
            ).scalars()
        )
        if not item_node_ids.issubset(approved_node_ids):
            raise HTTPException(400, "報帳只能列入已核准且已有額度的預算條目")
    entry.budget_included = body.included
    entry.budget_included_by_id = user_id
    entry.budget_included_at = datetime.now(UTC)
    if body.note:
        entry.note = body.note
    await db.flush()
    return entry


async def complete_expense_claim(db: AsyncSession, entry: JournalEntry) -> JournalEntry:
    if entry.source_type != "expense_claim":
        raise HTTPException(400, "只有報帳案件可以完成核銷")
    if entry.claim_status != ExpenseClaimStatus.APPROVED:
        raise HTTPException(400, "報帳尚未完成覆核")
    if entry.budget_included is not True:
        raise HTTPException(400, "報帳尚未列入已核准預算")
    if entry.payment_status in (None, ExpensePaymentStatus.UNPAID):
        raise HTTPException(400, "請先登錄付款或代墊償還")
    evidence_count = await db.scalar(
        select(func.count())
        .select_from(ExpenseClaimItemEvidence)
        .join(ExpenseClaimItem)
        .where(ExpenseClaimItem.journal_entry_id == entry.id)
    )
    if not entry.evidence_url and not evidence_count:
        raise HTTPException(400, "至少上傳一份憑證後才能完成核銷")
    entry.claim_status = ExpenseClaimStatus.COMPLETED
    await db.flush()
    return entry


async def create_transfer(
    db: AsyncSession, ledger_id: uuid.UUID, body: TransferCreate, user_id: uuid.UUID
) -> JournalEntry:
    if body.from_fund_account_id == body.to_fund_account_id:
        raise HTTPException(400, "轉出與轉入帳戶不得相同")
    funds = list(
        (
            await db.execute(
                select(FundAccount).where(
                    FundAccount.id.in_([body.from_fund_account_id, body.to_fund_account_id]),
                    FundAccount.ledger_id == ledger_id,
                    FundAccount.is_active,
                )
            )
        ).scalars()
    )  # noqa: E712
    if len(funds) != 2:
        raise HTTPException(400, "資金帳戶不存在或不屬於此帳本")
    lookup = {fund.id: fund for fund in funds}
    return await create_journal(
        db,
        ledger_id,
        JournalCreate(
            period_id=body.period_id,
            entry_date=body.entry_date,
            description=body.description,
            note=body.note,
            source_type="fund_transfer",
            source_event="transfer",
            lines=[
                {
                    "account_id": lookup[body.to_fund_account_id].chart_account_id,
                    "debit": body.amount,
                },
                {
                    "account_id": lookup[body.from_fund_account_id].chart_account_id,
                    "credit": body.amount,
                },
            ],
        ),
        user_id,
    )


async def create_expense_claim(
    db: AsyncSession, ledger_id: uuid.UUID, body: ExpenseClaimCreate, user_id: uuid.UUID
) -> JournalEntry:
    fund = await db.get(FundAccount, body.fund_account_id)
    if not fund or fund.ledger_id != ledger_id or not fund.is_active:
        raise HTTPException(400, "付款資金保管點不存在或已停用")
    expense_account = await db.get(ChartAccount, body.expense_account_id)
    if (
        not expense_account
        or expense_account.ledger_id != ledger_id
        or expense_account.account_type != FinanceAccountType.EXPENSE
        or not expense_account.is_active
    ):
        raise HTTPException(400, "支出科目不存在、非支出科目或已停用")
    amount = sum(
        _claim_item_total(item.unit_price, item.tax_rate, item.quantity) for item in body.items
    )
    payment_account_id = fund.chart_account_id
    if body.payment_method == ExpensePaymentMethod.ADVANCE:
        payable = await db.scalar(
            select(ChartAccount).where(
                ChartAccount.ledger_id == ledger_id,
                ChartAccount.account_type == FinanceAccountType.LIABILITY,
                ChartAccount.code == "2101",
            )
        )
        if not payable:
            raise HTTPException(400, "找不到代墊應付款科目")
        payment_account_id = payable.id
    for item in body.items:
        if item.budget_node_id:
            node = await db.get(FinanceBudgetNode, item.budget_node_id)
            if not node:
                raise HTTPException(400, "預算條目不存在")
            budget = await get_budget(db, node.budget_id)
            if budget.ledger_id != ledger_id or budget.period_id != body.period_id:
                raise HTTPException(400, "預算條目不屬於此帳本或會計期間")
            has_child = await db.scalar(
                select(FinanceBudgetNode.id)
                .where(FinanceBudgetNode.parent_id == node.id, FinanceBudgetNode.is_active)
                .limit(1)
            )
            if has_child:
                raise HTTPException(400, "報帳只能對應最末層預算條目")
        for evidence in item.evidence:
            validate_evidence_key(evidence.storage_key, ledger_id)
    entry = await create_journal(
        db,
        ledger_id,
        JournalCreate(
            period_id=body.period_id,
            entry_date=body.entry_date,
            description=f"報帳｜{body.description}（{len(body.items)} 項）",
            source_type="expense_claim",
            source_event="expense_claim",
            evidence_url=body.evidence_url,
            source_url=body.source_url,
            note=body.note,
            lines=[
                {"account_id": expense_account.id, "debit": amount},
                {"account_id": payment_account_id, "credit": amount},
            ],
        ),
        user_id,
        pending=True,
    )
    entry.claim_status = ExpenseClaimStatus.PENDING_REVIEW
    entry.procurement_status = ExpenseProcurementStatus.NOT_REQUIRED
    entry.payment_status = ExpensePaymentStatus.UNPAID
    entry.budget_included = False
    entry.proposing_org_id = body.proposing_org_id
    entry.advanced_by_id = body.advanced_by_id
    entry.payment_method = body.payment_method
    for item in body.items:
        claim_item = ExpenseClaimItem(
            journal_entry_id=entry.id,
            name=item.name,
            unit_price=item.unit_price,
            tax_rate=item.tax_rate,
            quantity=item.quantity,
            unit=item.unit,
            budget_node_id=item.budget_node_id,
            budget_exception_note=item.budget_exception_note,
        )
        db.add(claim_item)
        await db.flush()
        db.add_all(
            [
                ExpenseClaimItemEvidence(
                    item_id=claim_item.id,
                    storage_key=evidence.storage_key,
                    filename=evidence.filename,
                    content_type=evidence.content_type,
                    file_size=evidence.file_size,
                    evidence_type=evidence.evidence_type,
                    note=evidence.note,
                    uploaded_by_id=user_id,
                )
                for evidence in item.evidence
            ]
        )
    await db.flush()
    return entry


async def update_expense_claim(
    db: AsyncSession,
    entry: JournalEntry,
    body: ExpenseClaimCreate,
    user_id: uuid.UUID,
) -> JournalEntry:
    if entry.source_type != "expense_claim":
        raise HTTPException(400, "只有報帳案件可以修改")
    if entry.created_by_id != user_id:
        raise HTTPException(403, "只能修改自己提出的報帳")
    budgeted_reimbursement_edit = (
        entry.status == JournalStatus.POSTED
        and entry.claim_status == ExpenseClaimStatus.APPROVED
        and entry.budget_included is True
        and entry.payment_status in (None, ExpensePaymentStatus.UNPAID)
    )
    if not budgeted_reimbursement_edit and entry.status not in (
        JournalStatus.PENDING_REVIEW,
        JournalStatus.RETURNED,
    ):
        raise HTTPException(400, "只有待覆核、退回補正或已列入預算但尚未付款的報帳可以修改")
    if budgeted_reimbursement_edit and (
        body.period_id != entry.period_id
        or body.entry_date != entry.entry_date
        or body.payment_method != entry.payment_method
    ):
        raise HTTPException(400, "核銷階段只能調整報帳品項與金額")
    approved_budget_node_ids: set[uuid.UUID] = set()
    if budgeted_reimbursement_edit:
        item_node_ids = {item.budget_node_id for item in body.items}
        if None in item_node_ids or not item_node_ids:
            raise HTTPException(400, "進入核銷階段後，所有報帳品項都必須對應預算條目")
        approved_budget_node_ids = set(
            (
                await db.execute(
                    select(FinanceBudgetAllocation.node_id)
                    .join(FinanceBudgetSubmission)
                    .where(
                        FinanceBudgetAllocation.node_id.in_(item_node_ids),
                        FinanceBudgetSubmission.status == BudgetSubmissionStatus.APPROVED,
                    )
                )
            ).scalars()
        )
        if item_node_ids != approved_budget_node_ids:
            raise HTTPException(400, "進入核銷階段後，只能修改對應已核准預算的報帳品項")
    await validate_period(db, entry.ledger_id, body.period_id, body.entry_date)
    validate_evidence_key(body.evidence_url, entry.ledger_id)

    fund = await db.get(FundAccount, body.fund_account_id)
    if not fund or fund.ledger_id != entry.ledger_id or not fund.is_active:
        raise HTTPException(400, "付款資金保管點不存在或已停用")
    expense_account = await db.get(ChartAccount, body.expense_account_id)
    if (
        not expense_account
        or expense_account.ledger_id != entry.ledger_id
        or expense_account.account_type != FinanceAccountType.EXPENSE
        or not expense_account.is_active
    ):
        raise HTTPException(400, "支出科目不存在、非支出科目或已停用")
    amount = sum(
        _claim_item_total(item.unit_price, item.tax_rate, item.quantity) for item in body.items
    )
    payment_account_id = fund.chart_account_id
    if body.payment_method == ExpensePaymentMethod.ADVANCE:
        payable = await db.scalar(
            select(ChartAccount).where(
                ChartAccount.ledger_id == entry.ledger_id,
                ChartAccount.account_type == FinanceAccountType.LIABILITY,
                ChartAccount.code == "2101",
            )
        )
        if not payable:
            raise HTTPException(400, "找不到代墊應付款科目")
        payment_account_id = payable.id
    for item in body.items:
        if item.budget_node_id:
            node = await db.get(FinanceBudgetNode, item.budget_node_id)
            if not node:
                raise HTTPException(400, "預算條目不存在")
            budget = await get_budget(db, node.budget_id)
            if budget.ledger_id != entry.ledger_id or budget.period_id != body.period_id:
                raise HTTPException(400, "預算條目不屬於此帳本或會計期間")
            has_child = await db.scalar(
                select(FinanceBudgetNode.id)
                .where(FinanceBudgetNode.parent_id == node.id, FinanceBudgetNode.is_active)
                .limit(1)
            )
            if has_child:
                raise HTTPException(400, "報帳只能對應最末層預算條目")
        for evidence in item.evidence:
            validate_evidence_key(evidence.storage_key, entry.ledger_id)

    lines = list(
        (await db.execute(select(JournalLine).where(JournalLine.entry_id == entry.id))).scalars()
    )
    if len(lines) != 2:
        raise HTTPException(400, "報帳傳票明細格式不正確，無法修改")
    debit_line, credit_line = lines
    if debit_line.debit == 0 and credit_line.debit > 0:
        debit_line, credit_line = credit_line, debit_line
    if debit_line.debit == 0 or credit_line.credit == 0:
        raise HTTPException(400, "報帳傳票借貸明細不正確，無法修改")
    if budgeted_reimbursement_edit and (
        expense_account.id != debit_line.account_id
        or (
            entry.payment_method == ExpensePaymentMethod.DIRECT
            and fund.chart_account_id != credit_line.account_id
        )
    ):
        raise HTTPException(400, "核銷階段不可變更付款或支出科目")
    debit_line.account_id = expense_account.id
    debit_line.debit = amount
    debit_line.credit = 0
    credit_line.account_id = payment_account_id
    credit_line.debit = 0
    credit_line.credit = amount

    old_items = list(
        (
            await db.execute(
                select(ExpenseClaimItem).where(ExpenseClaimItem.journal_entry_id == entry.id)
            )
        ).scalars()
    )
    old_evidence = list(
        (
            await db.execute(
                select(ExpenseClaimItemEvidence).where(
                    ExpenseClaimItemEvidence.item_id.in_([item.id for item in old_items])
                )
            )
        ).scalars()
        if old_items
        else []
    )
    for evidence in old_evidence:
        await db.delete(evidence)
    for item in old_items:
        await db.delete(item)
    await db.flush()

    entry.period_id = body.period_id
    entry.entry_date = body.entry_date
    entry.description = f"報帳｜{body.description}（{len(body.items)} 項）"
    entry.source_url = body.source_url
    if body.evidence_url is not None:
        entry.evidence_url = body.evidence_url
    entry.note = body.note
    entry.proposing_org_id = body.proposing_org_id
    entry.advanced_by_id = body.advanced_by_id
    entry.payment_method = body.payment_method
    if not budgeted_reimbursement_edit:
        entry.payment_status = ExpensePaymentStatus.UNPAID
        entry.procurement_status = ExpenseProcurementStatus.NOT_REQUIRED
        entry.budget_included = False
        if entry.status == JournalStatus.RETURNED:
            entry.status = JournalStatus.PENDING_REVIEW
            entry.claim_status = ExpenseClaimStatus.PENDING_REVIEW
            entry.reviewed_by_id = None

    for item in body.items:
        claim_item = ExpenseClaimItem(
            journal_entry_id=entry.id,
            name=item.name,
            unit_price=item.unit_price,
            tax_rate=item.tax_rate,
            quantity=item.quantity,
            unit=item.unit,
            budget_node_id=item.budget_node_id,
            budget_exception_note=item.budget_exception_note,
        )
        db.add(claim_item)
        await db.flush()
        db.add_all(
            [
                ExpenseClaimItemEvidence(
                    item_id=claim_item.id,
                    storage_key=evidence.storage_key,
                    filename=evidence.filename,
                    content_type=evidence.content_type,
                    file_size=evidence.file_size,
                    evidence_type=evidence.evidence_type,
                    note=evidence.note,
                    uploaded_by_id=user_id,
                )
                for evidence in item.evidence
            ]
        )
    await db.flush()
    return entry


async def account_balances(db: AsyncSession, ledger_id: uuid.UUID) -> dict[uuid.UUID, int]:
    rows = (
        await db.execute(
            select(
                JournalLine.account_id,
                func.coalesce(func.sum(JournalLine.debit - JournalLine.credit), 0),
            )
            .join(JournalEntry)
            .where(JournalEntry.ledger_id == ledger_id, JournalEntry.status == JournalStatus.POSTED)
            .group_by(JournalLine.account_id)
        )
    ).all()
    return {row[0]: int(row[1]) for row in rows}


async def journal_with_lines(db: AsyncSession, entry: JournalEntry) -> dict:
    lines = list(
        (
            await db.execute(
                select(JournalLine, ChartAccount.name)
                .join(ChartAccount)
                .where(JournalLine.entry_id == entry.id)
            )
        ).all()
    )
    evidence_count = 0
    if entry.source_type == "expense_claim":
        evidence_count = int(
            await db.scalar(
                select(func.count())
                .select_from(ExpenseClaimItemEvidence)
                .join(ExpenseClaimItem)
                .where(ExpenseClaimItem.journal_entry_id == entry.id)
            )
            or 0
        )
    return {
        "id": entry.id,
        "ledger_id": entry.ledger_id,
        "period_id": entry.period_id,
        "entry_date": entry.entry_date,
        "description": entry.description,
        "status": entry.status,
        "created_by_id": entry.created_by_id,
        "reviewed_by_id": entry.reviewed_by_id,
        "posted_at": entry.posted_at,
        "source_type": entry.source_type,
        "source_id": entry.source_id,
        "source_event": entry.source_event,
        "source_url": entry.source_url,
        "evidence_url": entry.evidence_url,
        "note": entry.note,
        "claim_status": entry.claim_status,
        "procurement_status": entry.procurement_status,
        "procurement_updated_by_id": entry.procurement_updated_by_id,
        "procurement_updated_at": entry.procurement_updated_at,
        "payment_status": entry.payment_status,
        "payment_by_id": entry.payment_by_id,
        "payment_at": entry.payment_at,
        "budget_included": entry.budget_included,
        "budget_included_by_id": entry.budget_included_by_id,
        "budget_included_at": entry.budget_included_at,
        "proposing_org_id": entry.proposing_org_id,
        "advanced_by_id": entry.advanced_by_id,
        "payment_method": entry.payment_method,
        "reimbursement_entry_id": entry.reimbursement_entry_id,
        "evidence_complete": bool(entry.evidence_url) or evidence_count > 0,
        "lines": [
            {
                "id": line.id,
                "account_id": line.account_id,
                "debit": line.debit,
                "credit": line.credit,
                "memo": line.memo,
                "account_name": account_name,
            }
            for line, account_name in lines
        ],
    }


async def export_google_sheets(
    db: AsyncSession, ledger: FinanceLedger, spreadsheet_id: str
) -> None:
    """把已過帳資料同步到使用組織已授權的 Google 試算表。"""
    from api.models.google_calendar import OrgGoogleCalendarConfig
    from api.services.google_calendar_service import get_valid_credentials

    config = await db.scalar(
        select(OrgGoogleCalendarConfig).where(OrgGoogleCalendarConfig.org_id == ledger.org_id)
    )
    if not config or not config.is_connected:
        raise HTTPException(400, "請先在行事曆設定完成 Google 授權，再重新授權以加入 Sheets 權限")
    credentials = await get_valid_credentials(db, config)
    journals = list(
        (
            await db.execute(
                select(JournalEntry)
                .where(
                    JournalEntry.ledger_id == ledger.id, JournalEntry.status == JournalStatus.POSTED
                )
                .order_by(JournalEntry.entry_date, JournalEntry.created_at)
            )
        ).scalars()
    )
    values = [["日期", "摘要", "狀態", "來源", "借方", "貸方"]]
    for entry in journals:
        for line, account_name in (
            await db.execute(
                select(JournalLine, ChartAccount.name)
                .join(ChartAccount)
                .where(JournalLine.entry_id == entry.id)
            )
        ).all():
            values.append(
                [
                    str(entry.entry_date),
                    entry.description,
                    "已過帳",
                    account_name,
                    line.debit,
                    line.credit,
                ]
            )

    def _write() -> None:
        from googleapiclient.discovery import build

        sheets = build("sheets", "v4", credentials=credentials, cache_discovery=False)
        sheets.spreadsheets().values().clear(
            spreadsheetId=spreadsheet_id, range="財務總帳!A:F", body={}
        ).execute()
        sheets.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range="財務總帳!A1",
            valueInputOption="RAW",
            body={"values": values},
        ).execute()

    await asyncio.to_thread(_write)
